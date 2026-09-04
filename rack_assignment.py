import sys
import os
import time
import sys
from pathlib import Path
from openpyxl import load_workbook
from utils_misc import print_execution_seconds


# ============================================================
# Configuration
# ============================================================

EXCEL_FILE_NAME = "assignment.xlsx"

RACK_SHEET_NAME = "RackList"
TEAM_SHEET_NAME = "Team"
ASSIGNMENT_SHEET_NAME = "Assignment"


# ============================================================
# Utility functions
# ============================================================
def get_excel_file():
    if getattr(sys, "frozen", False):
        # Running as an EXE
        program_directory = Path(sys.executable).resolve().parent
    else:
        # Running as a Python script
        program_directory = Path(__file__).resolve().parent

    return program_directory / EXCEL_FILE_NAME


def is_empty(value):
    """
    Return True if an Excel cell is empty or contains
    only whitespace.
    """
    return value is None or str(value).strip() == ""


def normalize_task(value):
    """
    Convert the Other Tasks value to a normalized string.

    Examples:
        None       -> ""
        " PTO "    -> "pto"
        "Lead"     -> "lead"
        "Training" -> "training"
    """
    if is_empty(value):
        return ""

    return str(value).strip().lower()


# ============================================================
# Workbook functions
# ============================================================

def load_assignment_workbook(filename):
    """
    Load the Excel workbook.
    """
    return load_workbook(filename)


def validate_sheets(workbook):
    """
    Make sure the required sheets exist.
    """

    if RACK_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{RACK_SHEET_NAME}' was not found."
        )

    if TEAM_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{TEAM_SHEET_NAME}' was not found."
        )


# ============================================================
# Read RackList
# ============================================================

def read_racks(workbook):
    """
    Read rack names from column A of RackList.

    Row 1 is assumed to be the header.

    Returns:
        list[str]
    """

    sheet = workbook[RACK_SHEET_NAME]

    racks = []

    for row in sheet.iter_rows(
        min_row=2,
        max_col=1
    ):
        rack_name = row[0].value

        if not is_empty(rack_name):
            racks.append(str(rack_name).strip())

    return racks


# ============================================================
# Read Team
# ============================================================

def read_team(workbook):
    """
    Read technicians from the Team sheet.

    Column A = Technician
    Column B = Other Tasks

    Row 1 is assumed to be the header.

    Returns:
        list of dictionaries
    """

    sheet = workbook[TEAM_SHEET_NAME]

    technicians = []

    for row in sheet.iter_rows(
        min_row=2,
        max_col=2
    ):
        technician_name = row[0].value
        other_tasks = row[1].value

        # Ignore rows without a technician name
        if is_empty(technician_name):
            continue

        technicians.append({
            "name": str(technician_name).strip(),
            "other_tasks": (
                str(other_tasks).strip()
                if not is_empty(other_tasks)
                else ""
            ),
            "rack_count": 0,
            "racks": []
        })

    return technicians


# ============================================================
# Classify technicians
# ============================================================

def classify_technicians(technicians):
    """
    Divide technicians into four categories:

        PTO
        Lead
        Blank Other Tasks
        Other Tasks

    Returns:
        dictionary containing the four groups.
    """

    groups = {
        "pto": [],
        "lead": [],
        "blank": [],
        "other": []
    }

    for technician in technicians:

        task = normalize_task(
            technician["other_tasks"]
        )

        if task == "pto":
            groups["pto"].append(technician)

        elif task == "lead":
            groups["lead"].append(technician)

        elif task == "":
            groups["blank"].append(technician)

        else:
            groups["other"].append(technician)

    return groups


# ============================================================
# Calculate rack allocations
# ============================================================

def calculate_allocations(technicians, groups, total_racks):
    """
    Calculate how many racks each technician should receive.

    Rules:

        PTO:
            0 racks

        Lead:
            2 fewer racks than a normal technician,
            but at least 1 rack.

        Blank Other Tasks:
            1 more rack than a normal technician.

        Other Tasks:
            Normal allocation.

    If the racks cannot be divided evenly, the final
    normal technician may receive fewer racks.
    """

    blank_techs = groups["blank"]
    other_techs = groups["other"]
    lead_techs = groups["lead"]
    pto_techs = groups["pto"]

    # --------------------------------------------------------
    # PTO always gets zero
    # --------------------------------------------------------

    for technician in pto_techs:
        technician["rack_count"] = 0

    # --------------------------------------------------------
    # Technicians who can receive normal allocations
    # --------------------------------------------------------

    normal_techs = (
        blank_techs +
        other_techs
    )

    normal_count = len(normal_techs)
    lead_count = len(lead_techs)

    # --------------------------------------------------------
    # Special case: no normal technicians
    # --------------------------------------------------------

    if normal_count == 0:

        if lead_count == 0:

            if total_racks > 0:
                raise ValueError(
                    "There are racks available, but no "
                    "technicians can receive them."
                )

            return

        # Only Leads are available.
        # Give at least 1 rack to each Lead.
        if total_racks < lead_count:
            raise ValueError(
                "There are not enough racks to give "
                "each Lead at least 1 rack."
            )

        for technician in lead_techs:
            technician["rack_count"] = 1

        remaining = total_racks - lead_count

        # Any remaining racks go to the Leads
        index = 0

        while remaining > 0:

            lead_techs[
                index % lead_count
            ]["rack_count"] += 1

            remaining -= 1
            index += 1

        return

    # --------------------------------------------------------
    # Calculate the normal allocation.
    #
    # Let normal allocation = N
    #
    # Blank technician = N + 1
    # Other-task tech  = N
    # Lead              = N - 2
    #
    # We want Leads to have at least 1.
    #
    # Therefore N must be at least 3.
    # --------------------------------------------------------

    blank_count = len(blank_techs)
    other_count = len(other_techs)

    # Number of racks consumed by one "normal unit"
    active_count = (
        normal_count +
        lead_count
    )

    # Adjustment caused by:
    #
    # blank = +1
    # lead  = -2
    #
    adjustment = (
        blank_count -
        (lead_count * 2)
    )

    # Calculate possible normal allocation
    normal_allocation = (
        total_racks - adjustment
    ) // active_count

    # A Lead must have at least 1 rack
    if lead_count > 0:
        normal_allocation = max(
            normal_allocation,
            3
        )

    # --------------------------------------------------------
    # Assign initial allocations
    # --------------------------------------------------------

    for technician in blank_techs:
        technician["rack_count"] = (
            normal_allocation + 1
        )

    for technician in other_techs:
        technician["rack_count"] = (
            normal_allocation
        )

    for technician in lead_techs:
        technician["rack_count"] = max(
            normal_allocation - 2,
            1
        )

    # --------------------------------------------------------
    # Calculate how many racks remain.
    # --------------------------------------------------------

    allocated = sum(
        technician["rack_count"]
        for technician in technicians
    )

    remaining = total_racks - allocated

    # --------------------------------------------------------
    # Distribute extra racks ONLY to normal technicians.
    #
    # We deliberately don't use the Lead for this.
    # --------------------------------------------------------

    if remaining > 0:

        index = 0

        while remaining > 0:

            technician = normal_techs[
                index % len(normal_techs)
            ]

            technician["rack_count"] += 1

            remaining -= 1
            index += 1

    # --------------------------------------------------------
    # If we allocated too many because the minimum Lead
    # requirement was enforced, remove racks from normal
    # technicians.
    #
    # We remove from the end first, so the final normal
    # technician can have fewer racks.
    # --------------------------------------------------------

    elif remaining < 0:

        excess = abs(remaining)

        # Work backwards through normal technicians
        for technician in reversed(normal_techs):

            while (
                excess > 0
                and technician["rack_count"] > 0
            ):

                # Don't reduce a normal technician below zero
                technician["rack_count"] -= 1
                excess -= 1

            if excess == 0:
                break

        if excess > 0:
            raise ValueError(
                "Unable to satisfy the rack allocation "
                "rules with the available number of racks."
            )

    # --------------------------------------------------------
    # Final safety checks
    # --------------------------------------------------------

    for technician in lead_techs:

        if technician["rack_count"] < 1:

            raise RuntimeError(
                f"Lead '{technician['name']}' "
                f"received fewer than 1 rack."
            )

    final_count = sum(
        technician["rack_count"]
        for technician in technicians
    )

    if final_count != total_racks:

        raise RuntimeError(
            f"Rack allocation error. "
            f"Expected {total_racks}, "
            f"but allocated {final_count}."
        )

# ============================================================
# Assign actual rack names
# ============================================================

def assign_racks(technicians, rack_names):
    """
    Assign actual rack names to each technician.

    Rack names are taken sequentially from RackList.
    """

    rack_index = 0

    for technician in technicians:

        count = technician["rack_count"]

        technician["racks"] = rack_names[
            rack_index:
            rack_index + count
        ]

        rack_index += count

    # Safety check
    assigned_count = sum(
        len(technician["racks"])
        for technician in technicians
    )

    if assigned_count != len(rack_names):
        raise RuntimeError(
            f"Rack assignment error. "
            f"Expected {len(rack_names)} racks, "
            f"but assigned {assigned_count}."
        )


# ============================================================
# Create Assignment sheet
# ============================================================

def create_assignment_sheet(workbook):
    """
    Delete the existing Assignment sheet if it exists
    and create a new one.
    """

    if ASSIGNMENT_SHEET_NAME in workbook.sheetnames:
        del workbook[ASSIGNMENT_SHEET_NAME]

    return workbook.create_sheet(
        ASSIGNMENT_SHEET_NAME
    )


# ============================================================
# Write Assignment sheet
# ============================================================

def write_assignment_sheet(sheet, technicians):
    """
    Write technician assignments to the Assignment sheet.

    Each rack gets its own column.
    """

    # --------------------------------------------------------
    # Determine the maximum number of racks assigned to
    # any technician.
    # --------------------------------------------------------

    max_racks = max(
        (
            technician["rack_count"]
            for technician in technicians
        ),
        default=0
    )

    # --------------------------------------------------------
    # Create header row
    # --------------------------------------------------------

    headers = [
        "Technician",
        "Other Tasks",
        "Rack Count"
    ]

    for rack_number in range(1, max_racks + 1):
        headers.append(
            f"Rack {rack_number}"
        )

    sheet.append(headers)

    # --------------------------------------------------------
    # Write technicians
    # --------------------------------------------------------

    for technician in technicians:

        row = [
            technician["name"],
            technician["other_tasks"],
            technician["rack_count"]
        ]

        # Add each rack into its own column
        for rack_name in technician["racks"]:
            row.append(rack_name)

        sheet.append(row)


# ============================================================
# Format Assignment sheet
# ============================================================

def format_assignment_sheet(sheet):
    """
    Apply some basic formatting to the Assignment sheet.
    """

    # Freeze the header row
    sheet.freeze_panes = "A2"

    # Make headers bold
    for cell in sheet[1]:
        cell.font = cell.font.copy(
            bold=True
        )

    # Auto-size columns
    for column in sheet.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:

            if cell.value is not None:

                value_length = len(
                    str(cell.value)
                )

                if value_length > max_length:
                    max_length = value_length

        # Limit excessively wide columns
        sheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            40
        )


# ============================================================
# Print summary
# ============================================================

def print_summary(technicians, total_racks):
    """
    Print allocation information to the console.
    """

    print()
    print("=" * 70)
    print("RACK ASSIGNMENT SUMMARY")
    print("=" * 70)

    print(
        f"Total racks      : {total_racks}"
    )

    print(
        f"Technicians      : {len(technicians)}"
    )

    print("-" * 70)

    for technician in technicians:

        print(
            f"{technician['name']:20} "
            f"{technician['other_tasks']:15} "
            f"{technician['rack_count']:3} racks"
        )

    print("-" * 70)

    assigned = sum(
        technician["rack_count"]
        for technician in technicians
    )

    print(
        f"Total assigned   : {assigned}"
    )

    print("=" * 70)


# ============================================================
# Displays information about the program
# ============================================================
def show_startup_prompt():
    """
    Display information about what the program will do
    and the prerequisites before starting.
    """

    print()
    print("=" * 70)
    print("                RACK ASSIGNMENT PROGRAM")
    print("=" * 70)

    print()
    print("This program expects assignment.xlsx and will:")
    print()

    print("  1. Read rack names from the 'RackList' sheet.")
    print("  2. Read technicians and 'OTHER TASKS' from the 'Team' sheet.")
    print("  3. Allocate racks according to the following rules:")
    print("       - PTO  -> 0 racks")
    print("       - Lead -> 2 fewer racks than normal, minimum 1 rack")
    print("       - Blank OTHER TASKS -> 1 extra rack")
    print("       - has something for OTHER TASKS -> normal allocation")
    print("  4. Creates new sheet 'Assignment' with the rack allocation")
    print("  5. Save the results back into assignment.xlsx.")

    print()
    print("Prerequisites:")
    print()

    print(
        "  - assignment.xlsx must exist in the same folder as this program."
    )

    print(
        "  - 'RackList' sheet must exist."
    )

    print(
        "  - Rack names must be in column A, starting at row 2."
    )

    print(
        "  - 'Team' sheet must exist."
    )

    print(
        "  - Technician names must be in column A, starting at row 2."
    )

    print(
        "  - 'OTHER TASKS' must be in column B."
    )

    print(
        "  - Row 1 must contain the headers."
    )

    print()
    print("-" * 70)

    input("Press ENTER to continue...")

    print()


# ============================================================
# Main program
# ============================================================

def main():

    print("Loading workbook...")

    # --------------------------------------------------------
    # Load workbook
    # --------------------------------------------------------
    excel_file = get_excel_file()
    workbook = load_assignment_workbook(
        excel_file
    )

    # --------------------------------------------------------
    # Validate workbook
    # --------------------------------------------------------

    validate_sheets(workbook)

    # --------------------------------------------------------
    # Read data
    # --------------------------------------------------------

    rack_names = read_racks(workbook)

    technicians = read_team(workbook)

    if not rack_names:
        raise ValueError(
            "No racks were found in RackList."
        )

    if not technicians:
        raise ValueError(
            "No technicians were found in Team."
        )

    print(
        f"Found {len(rack_names)} racks."
    )

    print(
        f"Found {len(technicians)} technicians."
    )

    # --------------------------------------------------------
    # Classify technicians
    # --------------------------------------------------------

    groups = classify_technicians(
        technicians
    )

    # --------------------------------------------------------
    # Calculate allocations
    # --------------------------------------------------------

    calculate_allocations(
        technicians,
        groups,
        len(rack_names)
    )

    # --------------------------------------------------------
    # Assign actual rack names
    # --------------------------------------------------------

    assign_racks(
        technicians,
        rack_names
    )

    # --------------------------------------------------------
    # Create output sheet
    # --------------------------------------------------------

    assignment_sheet = create_assignment_sheet(
        workbook
    )

    # --------------------------------------------------------
    # Write results
    # --------------------------------------------------------

    write_assignment_sheet(
        assignment_sheet,
        technicians
    )

    # --------------------------------------------------------
    # Format results
    # --------------------------------------------------------

    format_assignment_sheet(
        assignment_sheet
    )

    # --------------------------------------------------------
    # Save to the SAME Excel file
    # --------------------------------------------------------

    workbook.save(
        excel_file
    )

    # --------------------------------------------------------
    # Display summary
    # --------------------------------------------------------

    print_summary(
        technicians,
        len(rack_names)
    )

    print()
    print(
        f"Assignment completed successfully."
    )

    print(
        f"Updated file: {excel_file}"
    )

# ============================================================
# Program entry point
# ============================================================

if __name__ == "__main__":
    print ("Executing rack assignment script...")
    show_startup_prompt()
    start_time = time.time()
    main()
    end_time = time.time()

    print("\nADAM base parser completed successfully!" )
    print_execution_seconds(start_time, end_time)
