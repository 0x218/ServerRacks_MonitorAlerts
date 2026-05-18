# ============================================================
# Project : Server Rack Monitoring & Alert System
# File    : ibmnon_base_parser.py
# Author  : Renjith Kumar
# Created : 
# Purpose : 
#
# Description:
#
#
# ============================================================

import os
import time
from datetime import datetime
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from utils_ibmrack import (
    wait,
    create_driver,
    login,
    navigate_to_page,
    expand_pod_if_collapsed,
)

from utils_config import (
    load_config
)


sheet_name = "Minimal"


# =========================================================
# EXTRACT RACK
# =========================================================
def extract_rack_data(space_element):
    try:
        rack_numbers_elements = space_element.find_elements(
            By.CLASS_NAME,
            "rack-numbers"
        )
        if not rack_numbers_elements:
            return None
        
        # SPACE LABEL
        space_label = space_element.find_element(
            By.CLASS_NAME,
            "space-label"
        ).text.strip()

        # RACK LETTERS
        rack_letters = space_element.find_element(
            By.CLASS_NAME,
            "rack-letters"
        ).text.strip()

        # RACK NUMBERS
        # rack_numbers = rack_numbers_elements[0].text.strip()
        rack_numbers = space_element.find_element(
            By.CLASS_NAME,
            "rack-numbers"
        ).text.strip()

        # RACK MODEL
        rack_model = space_element.find_element(
            By.CLASS_NAME,
            "rack-model"
        ).text.strip()

        # PROGRESS %
        progress_percentage = space_element.find_element(
            By.CLASS_NAME,
            "progress-percentage"
        ).text.strip()

        # KEEP ONLY ZORA01
        if rack_model.upper() != "ZORA01":
            return None

        # SERIAL
        rack_serial = f"{rack_letters}{rack_numbers}"

        return [
            space_label,
            rack_serial,
            progress_percentage
        ]

    except Exception as e:
        print("Failed parsing rack:", e)
        return None


# =========================================================
# PARSE SPACES
# =========================================================
def parse_spaces_in_pod(
        pod_element,
        prefixes_for_this_pod
):

    results = []

    spaces = pod_element.find_elements(
        By.CLASS_NAME,
        "space"
    )

    print(
        f"Scanning {len(spaces)} spaces "
        f"for prefixes: {prefixes_for_this_pod}"
    )

    for space in spaces:
        try:
            space_label = space.find_element(
                By.CLASS_NAME,
                "space-label"
            ).text.strip().upper()

            # EXAMPLE:
            # 6V401 startswith 6V4

            if any(
                    space_label.startswith(prefix)
                    for prefix in prefixes_for_this_pod
            ):
                rack_data = extract_rack_data(space)
                if rack_data:
                    results.append(rack_data)
                    print(
                        f"Matched Rack: "
                        f"{space_label}"
                    )
        except Exception:
            continue

    return results

# =========================================================
# EXCEL
# =========================================================
def append_results_to_excel(
        results,
        save_folder
):
    os.makedirs(save_folder, exist_ok=True)
    date_str = datetime.now().strftime(
        "%m%d%Y"
    )

    file_name = (
        f"IBMRACK-Parser-"
        f"{date_str}.xlsx"
    )

    excel_path = os.path.join(
        save_folder,
        file_name
    )


    # CURRENT TIME COLUMN HEADER
    #
    # Example:
    # 01:48 AM
    current_time_header = datetime.now().strftime(
        "%I:%M %p"
    )

    # LOAD OR CREATE WORKBOOK
    if os.path.exists(excel_path):
        workbook = load_workbook(
            excel_path
        )
    else:
        workbook = Workbook()

    # GET OR CREATE specific SHEET
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # REMOVE DEFAULT SHEET
        if "Sheet" in workbook.sheetnames:
            default_sheet = workbook["Sheet"]
            workbook.remove(default_sheet)
        # CREATE specific SHEET
        sheet = workbook.create_sheet(
            title=sheet_name
        )

        # INITIAL HEADERS
        sheet.cell(
            row=1,
            column=1
        ).value = "Pod Location"

        sheet.cell(
            row=1,
            column=2
        ).value = "Rack Serial"
    # FIND / CREATE TIME COLUMN
    max_column = sheet.max_column
    time_column = None

    for col in range(1, max_column + 1):
        header_value = sheet.cell(
            row=1,
            column=col
        ).value

        if header_value == current_time_header:

            time_column = col
            break

    # CREATE NEW TIME COLUMN
    if time_column is None:
        time_column = max_column + 1
        sheet.cell(
            row=1,
            column=time_column
        ).value = current_time_header

    # BUILD EXISTING RACK MAP
    # KEY: (pod_location, rack_serial)
    existing_rows = {}
    max_row = sheet.max_row

    for row in range(2, max_row + 1):
        pod_location = sheet.cell(
            row=row,
            column=1
        ).value

        rack_serial = sheet.cell(
            row=row,
            column=2
        ).value

        if pod_location and rack_serial:
            key = (
                str(pod_location).strip(),
                str(rack_serial).strip()
            )

            existing_rows[key] = row

    # WRITE DATA
    for result in results:
        pod_location = result[0]
        rack_serial = result[1]
        progress_percentage = result[2]

        # KEY:
        # (pod_location, rack_serial)
        key = (
            str(pod_location).strip(),
            str(rack_serial).strip()
        )

        # EXISTING ROW
        if key in existing_rows:
            row_number = existing_rows[key]

        # NEW ROW
        else:
            row_number = sheet.max_row + 1
            existing_rows[key] = row_number

        # ALWAYS UPDATE LOCATION
        sheet.cell(
            row=row_number,
            column=1
        ).value = pod_location

        # ALWAYS UPDATE SERIAL
        sheet.cell(
            row=row_number,
            column=2
        ).value = rack_serial

        # WRITE %
        sheet.cell(
            row=row_number,
            column=time_column
        ).value = progress_percentage

    # SAVE
    workbook.save(excel_path)
    print(
        f"\nUpdated Excel: {excel_path}"
    )


# =========================================================
# PROCESS PODS
# =========================================================
def process_pods(
        driver,
        rack_prefixes,
        save_folder
):
    all_results = []

    # BUILD POD MAP
    # Example:
    #
    # {
    #   "6V": ["6V4"],
    #   "6W": ["6W1"]
    # }
    #
    pod_map = {}

    for prefix in rack_prefixes:
        pod_code = prefix[:2]
        if pod_code not in pod_map:
            pod_map[pod_code] = []

        pod_map[pod_code].append(prefix)

    print("\nPod Map:")
    print(pod_map)

    # PROCESS EACH POD
    for pod_code, prefixes_for_this_pod in pod_map.items():
        print(f"\nProcessing Pod {pod_code}")
        print(
            f"Using prefixes: "
            f"{prefixes_for_this_pod}"
        )

        try:
            # EXPAND POD
            pod = expand_pod_if_collapsed(
                driver,
                pod_code
            )

            # PARSE SPACES
            pod_results = parse_spaces_in_pod(
                pod,
                prefixes_for_this_pod
            )

            print(
                f"Found "
                f"{len(pod_results)} "
                f"matching racks"
            )

            all_results.extend(
                pod_results
            )

        except Exception as e:
            print(
                f"Failed processing "
                f"Pod {pod_code}: {e}"
            )

    # SAVE RESULTS
    if all_results:
        append_results_to_excel(
            all_results,
            save_folder
        )

    else:
        print("\nNo matching racks found")


# =========================================================
# MAIN
# =========================================================
def main():
    print("Running IBMRACK Base Parser...")
    settings = load_config()
    driver = create_driver(
        settings["driver_path"]
    )

    try:
        login(
            driver,
            settings["base_url"],
            settings["username"],
            settings["password"]
        )

        print(
            "Successfully logged in"
        )

        wait(2)

        navigate_to_page(
            driver,
            settings["base_url"],
            "/module/mapv2/map.php?b=1flex"
        )

        driver.maximize_window()

        print(
            "Successfully navigated "
            "to page"
        )

        # RACK PREFIXES
        rack_prefixes = settings[
            "rack_prefixes"
        ]

        print(
            "\nConfigured Prefixes:"
        )

        print(rack_prefixes)

        # PROCESS
        process_pods(
            driver,
            rack_prefixes,
            settings["save_folder"]
        )

        wait(10)

    finally:
        driver.quit()


# =========================================================
# START
# =========================================================
if __name__ == "__main__":
    start_time = time.time()

    main()

    end_time = time.time()

    execution_seconds = (
        end_time - start_time
    )

    execution_minutes = (
        execution_seconds / 60
    )

    print("\nIBMRACK base parser completed successfully!" )

    print(
        f"Total Execution Time: "
        f"{execution_seconds:.2f} seconds "
        f"({execution_minutes:.2f} minutes)"
    )
