# ============================================================
# Project : Server Rack Monitoring & Alert System
# File    : ibmnon_deep_parser.py
# Author  : Renjith Kumar
# Created : 
# Purpose : 
#
# Description:
#
#
# ============================================================
import os
import time
from datetime import datetime
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from utils_ibmrack import (
    wait,
    get_network_status,
    get_power_status,
    create_driver,
    login,
    navigate_to_page,
    expand_pod_if_collapsed,
    get_tooltip,
    close_tooltip,
    parse_tooltip_header,
)

from utils_config import (
    load_config
)

sheet_name = "Detailed"

# =========================================================
# SAVE HTML
# =========================================================
def save_page_html(driver, save_folder, file_name):
    os.makedirs(save_folder, exist_ok=True)
    save_path = os.path.join(save_folder, file_name)
    html = driver.page_source

    with open(save_path, "w", encoding="utf-8") as f:
        f.write(html)

    print("Saved HTML to:", save_path)
    return save_path

# =========================================================
# FIND ZORA RACKS
# =========================================================
def find_zora_racks(pod_container, rack_prefixes):

    matching_spaces = []

    spaces = pod_container.find_elements(
        By.CSS_SELECTOR,
        "div.space.flex.occupied"
    )

    for space in spaces:
        try:
            # SPACE LABEL
            space_label = space.find_element(
                By.CLASS_NAME,
                "space-label"
            ).text.strip()

            # PREFIX FILTER
            if not any(
                    space_label.startswith(prefix)
                    for prefix in rack_prefixes
            ):
                continue

            # MODEL
            rack_model = space.find_element(
                By.CLASS_NAME,
                "rack-model"
            ).text.strip()

            # PROGRESS %
            rack_progress_percentage = space.find_element(
                By.CLASS_NAME,
                "progress-percentage"
            ).text.strip()

            if rack_model.upper() != "ZORA01":
                continue

            matching_spaces.append({
                "space": space,
                "rack_progress_percentage": rack_progress_percentage
            })

        except Exception:
            continue

    return matching_spaces


# OPEN TOOLTIP
def open_tooltip(driver, space):
    # CLOSE OLD TOOLTIP
    try:
        close_buttons = driver.find_elements(
            By.CLASS_NAME,
            "tooltip-close"
        )

        if close_buttons:
            driver.execute_script(
                "arguments[0].click();",
                close_buttons[0]
            )

            time.sleep(1)
    except:
        pass

    # SCROLL ELEMENT INTO VIEW
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center'});",
        space
    )

    time.sleep(1)

    # CLICK USING JAVASCRIPT
    driver.execute_script(
        "arguments[0].click();",
        space
    )
    time.sleep(2)


# =========================================================
# PARSE TOOLTIP TABLE
# =========================================================
def parse_tooltip_table(tooltip):
    parsed_rows = []
    table_rows = tooltip.find_elements(
        By.CSS_SELECTOR,
        "tbody tr.uut-row"
    )

    for row in table_rows:
        uut_type = row.get_attribute(
            "data-uut-type"
        )

        if uut_type != "SERVER":
            continue

        tds = row.find_elements(By.TAG_NAME, "td")

        if len(tds) < 12:
            continue

        try:
            loc_value = tds[0].text.strip()
            type_value = tds[1].text.strip()
            serial_link = tds[2].find_element(
                By.TAG_NAME,
                "a"
            )

            serial_value = serial_link.text.strip()

            # NETWORK
            network_td = tds[3]

            network_dot = network_td.find_element(
                By.CLASS_NAME,
                "status-dot"
            )

            network_status = get_network_status(network_dot)

            # POWER
            power_td = tds[4]
            power_status = get_power_status(power_td)

            # STATUS
            status_value = tds[9].text.strip()
            station_value = tds[10].text.strip()
            idle_time_value = tds[11].text.strip()

            row_data = {
                "loc": loc_value,
                "type": type_value,
                "ct_serial_number": serial_value,
                "network": network_status,
                "power": power_status,
                "status": status_value,
                "station": station_value,
                "idle_time": idle_time_value
            }

            parsed_rows.append(row_data)

        except Exception as e:
            print(f"Failed to parse row: {e}")

    return parsed_rows


# =========================================================
# EXCEL FILE
# =========================================================
def get_excel_file(excel_save_folder):
    save_folder = excel_save_folder

    os.makedirs(save_folder, exist_ok=True)

    date_str = datetime.now().strftime("%m%d%Y")
    file_name = (
        f"IBMRACK-Parser-"
        f"{date_str}.xlsx"
    )

    excel_path = os.path.join(
        save_folder,
        file_name
    )

    return excel_path



# =========================================================
# WRITE TO EXCEL
# =========================================================
def write_to_excel(
        excel_save_folder,
        rack_progress_percentage,
        header_data,
        table_rows
):
    # CREATE SAVE FOLDER
    os.makedirs(
        excel_save_folder,
        exist_ok=True
    )

    # EXCEL FILE NAME
    date_str = datetime.now().strftime(
        "%m%d%Y"
    )

    file_name = (
        f"IBMRACK-Parser-"
        f"{date_str}.xlsx"
    )

    excel_path = os.path.join(
        excel_save_folder,
        file_name
    )

    # LOAD OR CREATE WORKBOOK
    if os.path.exists(excel_path):
        workbook = load_workbook(
            excel_path
        )
    else:
        workbook = Workbook()

    # GET OR CREATE specific SHEET
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # REMOVE DEFAULT SHEET
        if "Sheet" in workbook.sheetnames:
            default_sheet = workbook["Sheet"]
            workbook.remove(default_sheet)
        # CREATE specific SHEET
        sheet = workbook.create_sheet(
            title=sheet_name
        )

        headers = [
            "Current Time",
            "Rack Location",
            "Rack Serial#",
            "Rack Progress %",
            "Time In Test",
            "LOC",
            "Type",
            "CT Serial#",
            "Network",
            "Power",
            "Status",
            "Station",
            "Idle Since"
        ]
        sheet.append(headers)

    # CURRENT TIME
    current_time = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    # WRITE SERVER ROWS
    for server_row in table_rows:
        excel_row = [
            current_time,
            # HEADER DATA
            header_data["location"],
            header_data["rack_serial_number"],

            # RACK %
            rack_progress_percentage,

            # TIME IN TEST
            header_data["time_in_test"],

            # SERVER DATA
            server_row["loc"],
            server_row["type"],
            server_row["ct_serial_number"],
            server_row["network"],
            server_row["power"],
            server_row["status"],
            server_row["station"],
            server_row["idle_time"]
        ]
        sheet.append(excel_row)

    # SAVE
    workbook.save(excel_path)

    print(f"Wrote {len(table_rows)} rows to {sheet_name}")

# =========================================================
# PROCESS POD
# =========================================================
def process_pod(
        driver,
        pod_name,
        rack_prefixes,
        excel_save_folder
):
    pod_container = expand_pod_if_collapsed(
        driver,
        pod_name
    )

    matching_spaces = find_zora_racks(
        pod_container,
        rack_prefixes
    )

    print(
        f"Found {len(matching_spaces)} matching racks"
    )

    for rack_data  in matching_spaces:
        space = rack_data["space"]

        rack_progress_percentage = rack_data[
            "rack_progress_percentage"
        ]
        try:
            open_tooltip(driver, space)
            tooltip = get_tooltip(driver)
            header_data = parse_tooltip_header(
                tooltip
            )

            # Wait for tooltip live data to load
            time.sleep(15)

            table_rows = parse_tooltip_table(
                tooltip
            )

            write_to_excel(
                excel_save_folder,
                rack_progress_percentage,
                header_data,
                table_rows
            )

            close_tooltip(driver)

        except Exception as e:
            print(f"Failed processing rack: {e}")


# =========================================================
# MAIN
# =========================================================
def main():
    print("Running IBMRACK Deep Parser...")
    settings = load_config()

    driver = create_driver(
        settings["driver_path"]
    )

    try:

        login(
            driver,
            settings["base_url"],
            settings["username"],
            settings["password"]
        )

        wait(2)

        navigate_to_page(
            driver,
            settings["base_url"],
            "/module/mapv2/map.php?b=1flex"
        )

        wait(3)

        # Save the page for offline access
        # save_page_html(
        #    driver,
        #    settings["save_folder"],
        #    "map_page.html"
        # )

        # PODS
        pod_names = settings["pod_names"]

        # PREFIX FILTERS
        rack_prefixes = settings["rack_prefixes"]

        # PROCESS
        for pod_name in pod_names:
            print(f"\nProcessing {pod_name}")

            process_pod(
                driver,
                pod_name,
                rack_prefixes,
                settings["save_folder"]
            )

        print("\nCompleted" )
        wait(5)
    finally:
        driver.quit()


# =========================================================
# START
# =========================================================
if __name__ == "__main__":
    start_time = time.time()

    main()

    end_time = time.time()
    execution_seconds = end_time - start_time
    execution_minutes = execution_seconds / 60
    print("\nIBMRACK deep parser completed successfully!" )

    print(
        f"Total Execution Time: "
        f"{execution_seconds:.2f} seconds "
        f"({execution_minutes:.2f} minutes)"
    )