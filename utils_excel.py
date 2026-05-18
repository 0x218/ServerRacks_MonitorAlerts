import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


def get_daily_workbook_path(save_folder):
    os.makedirs(save_folder, exist_ok=True)
    today = datetime.now().strftime("%m%d%Y")
    file_name = f"IBMRACK-Parser-{today}.xlsx"

    return os.path.join(
        save_folder,
        file_name
    )


def get_or_create_sheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    sheet = workbook.create_sheet(title=sheet_name)
    return sheet


def open_daily_workbook(save_folder):
    excel_path = get_daily_workbook_path(
        save_folder
    )

    if os.path.exists(excel_path):
        workbook = load_workbook(
            excel_path
        )
    else:
        workbook = Workbook()
        # REMOVE DEFAULT SHEET
        default_sheet = workbook.active
        workbook.remove(default_sheet)
    return workbook, excel_path


def save_workbook(workbook, excel_path):
    workbook.save(excel_path)

